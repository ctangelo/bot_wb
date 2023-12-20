import openpyxl as op
import pandas as pd
from openpyxl.styles import Font, Alignment, Border, borders, PatternFill
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles.borders import Border
from pandas.io.excel import ExcelWriter
from openpyxl import load_workbook

from openpyxl.drawing.image import Image
import os
import numpy as np



def gen_analitica(file_2, file_1, user_id):

    download_path = f"/root/doc/{user_id}/file_1.xlsx"
    middle_path = f"/root/doc/{user_id}/Отчет.xlsx"

    # Функция для записи
    def update_spreadsheet(path: str, _df, startcol: int = 1, startrow: int = 1, sheet_name: str = "sheet1"):
        wb = op.load_workbook(path)
        for ir in range(0, len(_df)):
            for ic in range(0, len(_df.iloc[ir])):
                wb[sheet_name].cell(startrow + ir, startcol + ic).value = _df.iloc[ir][ic]
        wb.save(path)

    pd.set_option('max_colwidth', 120)
    pd.set_option('display.width', 500)


    df = pd.read_excel(file_1)
    df2 = pd.read_excel(file_2, sheet_name='Sheet1')
    df1_1 = pd.read_excel(file_1, sheet_name='Товары', skiprows=1)


    code = df2['Код номенклатуры'].unique()
    # articul = np.sort(df2['Артикул поставщика'].unique())
    articul = (df2['Артикул поставщика'].unique())


    list1 = []
    list2 = []

    for i in code:
        list1.append(next(iter(df2.loc[(df2['Код номенклатуры'] == i), 'Артикул поставщика']), 'no match'))

        list1.append(df2.loc[(df2['Код номенклатуры'] == i) & (df2['Тип документа'] == 'Возврат'), 'Вайлдберриз реализовал Товар (Пр)'].sum())
        list1.append(df2.loc[(df2['Код номенклатуры'] == i) & (df2['Тип документа'] == 'Продажа'), 'Вайлдберриз реализовал Товар (Пр)'].sum())

        list1.append(df2.loc[(df2['Код номенклатуры'] == i) & (df2['Тип документа'] == 'Возврат'), 'Вознаграждение с продаж до вычета услуг поверенного, без НДС'].sum())
        list1.append(df2.loc[(df2['Код номенклатуры'] == i) & (df2['Тип документа'] == 'Продажа'), 'Вознаграждение с продаж до вычета услуг поверенного, без НДС'].sum())

        list1.append(df2.loc[(df2['Код номенклатуры'] == i) & (df2['Тип документа'] == 'Возврат'), 'Доплаты'].sum())
        list1.append(df2.loc[(df2['Код номенклатуры'] == i) & (df2['Тип документа'] == 'Продажа'), 'Доплаты'].sum())
        
        list1.append(df2.loc[(df2['Код номенклатуры'] == i) & (df2['Тип документа'] == 'Возврат'), 'К перечислению Продавцу за реализованный Товар'].sum())
        list1.append(df2.loc[(df2['Код номенклатуры'] == i) & (df2['Тип документа'] == 'Продажа'), 'К перечислению Продавцу за реализованный Товар'].sum())
        
        list1.append(df2.loc[(df2['Код номенклатуры'] == i) & (df2['Тип документа'] == 'Возврат'), 'Кол-во'].sum())
        list1.append(df2.loc[(df2['Код номенклатуры'] == i) & (df2['Тип документа'] == 'Продажа'), 'Кол-во'].sum())

        list1.append(df2.loc[(df2['Код номенклатуры'] == i) & (df2['Тип документа'] == 'Возврат'), 'Количество доставок'].sum())
        list1.append(df2.loc[(df2['Код номенклатуры'] == i) & (df2['Тип документа'] == 'Продажа'), 'Количество доставок'].sum())

        list1.append(df2.loc[(df2['Код номенклатуры'] == i) & (df2['Тип документа'] == 'Возврат'), 'НДС с Вознаграждения Вайлдберриз'].sum())
        list1.append(df2.loc[(df2['Код номенклатуры'] == i) & (df2['Тип документа'] == 'Продажа'), 'НДС с Вознаграждения Вайлдберриз'].sum())

        list1.append(df2.loc[(df2['Код номенклатуры'] == i) & (df2['Тип документа'] == 'Возврат'), 'Общая сумма штрафов'].sum())
        list1.append(df2.loc[(df2['Код номенклатуры'] == i) & (df2['Тип документа'] == 'Продажа'), 'Общая сумма штрафов'].sum())

        list1.append("")
        list1.append(df2.loc[(df2['Код номенклатуры'] == i), 'Услуги по доставке товара покупателю'].sum())

        list1.append(df2.loc[(df2['Код номенклатуры'] == i) & (df2['Тип документа'] == 'Возврат'), 'Цена розничная с учетом согласованной скидки'].sum())
        list1.append(df2.loc[(df2['Код номенклатуры'] == i) & (df2['Тип документа'] == 'Продажа'), 'Цена розничная с учетом согласованной скидки'].sum())

        list2.append(list1)
        list1 = []

    list2.append([' ', 'Возврат', 'Продажа', 'Возврат', 'Продажа', 'Возврат', 'Продажа', 'Возврат', 'Продажа', 'Возврат', 'Продажа', 'Возврат', 
                  'Продажа', 'Возврат', 'Продажа', 'Возврат', 'Продажа', 'Возврат', 'Продажа', 'Возврат', 'Продажа'])

    list2.sort()

    fullstats_svodnaya = pd.DataFrame(list2, columns = ['Артикул поставщика', 'Вайлдберриз реализовал Товар (Пр)', 'Продажа', 
                                                        'Вознаграждение с продаж до вычета услуг поверенного, без НДС	', 'Продажа', 'Доплаты', 'Продажа', 
                                                        'К перечислению Продавцу за реализованный Товар	', 'Продажа', 'Кол-во', 'Продажа', 'Количество доставок', 
                                                        'Продажа', 'НДС с Вознаграждения Вайлдберриз	', 'Продажа', 'Общая сумма штрафов	', 'Продажа', 
                                                        'Услуги по доставке товара покупателю	', 'Продажа', '', 'Продажа'])

    list2 = []


    # делаем Себстоимость и Затраты на РК
    for i in articul:
        list1.append(i)
        list1.append(next(iter(df2.loc[(df2['Артикул поставщика'] == i), 'Код номенклатуры']), 'no match'))
        list2.append(list1)
        list1 = []

    marketing_df = pd.DataFrame(list2, columns=['Артикул поставщика', 'Номенклатура'])
    marketing_df['Затраты на Р/К'] = ''
    price_df = pd.DataFrame(list2, columns=['Артикул поставщика', 'Номенклатура'])
    price_df['Себестоимость'] = ''


    list2 = []
    for i in articul:
        list1.append(i)
        list1.append(len(df2[(df2['Артикул поставщика'] == i) & (df2['Обоснование для оплаты'] == 'Продажа')]))
        list2.append(list1)
        list1 = []


    # делаем Продажи
    sales_df = pd.DataFrame(list2, columns=['Артикул поставщика', 'Продажа'])


    # Аналитика 
    fullstats_analytica = []
    fullstats_analytica2 = []
    k = 0
    fullstats_analytica.append(['Бренд', 'Категория',	'Артикул поставщика', 'Номенклатура', 'Себестоимость', 'Кол-во Заказов', 'Количество продаж ( с учетом возврат и отмен)', 
                                'Цена розничная с учетом согласованной скидки', 'Вайлдберриз реализовал Товар (Пр)','Комиссия WB (до спп)', 'Комиссия WB (после спп)', 
                                'Цена', 'К перечислению Продавцу за реализованный', 'Логистика', 'Логистика средняя', 'Общая сумма штрафов', 'Себестоимость сумма', 'Налог', 
                                'Выручка (минус все расходов мп)', 'Маржинальность', 'ROI %', 'ЧП по SKU', 'Остатки на складе, шт', 'Остатки МП, шт', 
                                'Среднее кол-во заказов в день, шт', 'Остаток денег (wb)', 'Остаток денег (свой склад)', 'Процент выкупа, %', 'Конверсия в корзину, %', 
                                'Конверсия в заказ, %', 'Закончится через WB', 'Закончится через wb  / (fbs)', 'Р/К бюджет', 'ДРР'])
    

    for i in articul:
        fullstats_analytica2.append(next(iter(df2.loc[(df2['Артикул поставщика'] == i), 'Бренд']), 'no match'))
        fullstats_analytica2.append(next(iter(df2.loc[(df2['Артикул поставщика'] == i), 'Предмет']), 'no match'))
        fullstats_analytica2.append(i)
        fullstats_analytica2.append((next(iter(df2.loc[(df2['Артикул поставщика'] == i), 'Код номенклатуры']), 'no match')))
        fullstats_analytica2.append(f"=VLOOKUP(C{39+k},'Себестоимость'!A:C,3,FALSE)")
        fullstats_analytica2.append(f"='Сводная таблица'!K{3+k}-'Сводная таблица'!J{3+k}")
        fullstats_analytica2.append(f"=IFERROR(VLOOKUP(C{39+k},'Продажи'!A:B,2,FALSE),0)")
        fullstats_analytica2.append(f"='Сводная таблица'!U{3+k}-'Сводная таблица'!T{3+k}")
        fullstats_analytica2.append(f"='Сводная таблица'!C{3+k}-'Сводная таблица'!B{3+k}")
        fullstats_analytica2.append(f"=H{39+k}-M{39+k}")
        fullstats_analytica2.append(f"=('Сводная таблица'!E{3+k}+'Сводная таблица'!O{3+k})-('Сводная таблица'!D{3+k}+'Сводная таблица'!N{3+k})")
        fullstats_analytica2.append(f"=IF(G{39+k}=0,0,M{39+k}/G{39+k})")
        fullstats_analytica2.append(f"=('Сводная таблица'!I{3+k}-'Сводная таблица'!H{3+k})")
        fullstats_analytica2.append(f"='Сводная таблица'!S{3+k}")
        fullstats_analytica2.append(f"=IF(F{39+k}=0,0,('Сводная таблица'!S{3+k}/F{39+k}))")
        fullstats_analytica2.append(f"='Сводная таблица'!Q{3+k} -'Сводная таблица'!P{3+k}")
        fullstats_analytica2.append(f"=E{39+k}*G{39+k}")
        fullstats_analytica2.append(f"=I{39+k}*$B$23/100")
        fullstats_analytica2.append(f"=M{39+k}-N{39+k}-P{39+k}")
        fullstats_analytica2.append(f"=IF(S{39+k}=0,0,V{39+k}/S{39+k})")
        fullstats_analytica2.append(f"=IF(Q{39+k}=0,0,V{39+k}/Q{39+k})")
        fullstats_analytica2.append(f"=M{39+k}-N{39+k}-P{39+k}-Q{39+k}-R{39+k}-AG{39+k}")
        fullstats_analytica2.append((df1_1.loc[(df1_1['Артикул продавца'] == i), 'Остатки склад, шт']).sum())
        fullstats_analytica2.append((df1_1.loc[(df1_1['Артикул продавца'] == i), 'Остатки МП, шт']).sum())
        fullstats_analytica2.append((df1_1.loc[(df1_1['Артикул продавца'] == i), 'Среднее количество заказов в день, шт']).sum())
        fullstats_analytica2.append(f"=E{39+k}*W{39+k}")
        fullstats_analytica2.append(f"=X{39+k}*E{39+k}")
        fullstats_analytica2.append((df1_1.loc[(df1_1['Артикул продавца'] == i), 'Процент выкупа']).sum())
        fullstats_analytica2.append((df1_1.loc[(df1_1['Артикул продавца'] == i), 'Конверсия в корзину, %']).sum())
        fullstats_analytica2.append((df1_1.loc[(df1_1['Артикул продавца'] == i), 'Конверсия в заказ, %']).sum())
        fullstats_analytica2.append(f"=IF(Y{39+k}=0,0,W{39+k}/Y{39+k})")
        fullstats_analytica2.append(f"=IF(Y{39+k}=0,0,X{39+k}/Y{39+k})")
        fullstats_analytica2.append(f"=IFERROR(VLOOKUP(C{39+k},'Затраты на РК'!A:C,3,FALSE),0)")
        fullstats_analytica2.append(f"=IF(S{39+k}=0,0,AG{39+k}/S{39+k})")

        k += 1
        
        fullstats_analytica.append(fullstats_analytica2)
        fullstats_analytica2 = []


    fullstats1_df = pd.DataFrame(fullstats_analytica, columns=['Бренд', 'Категория',	'Артикул поставщика', 'Номенклатура', 'Себестоимость', 'Кол-во Заказов', 
                                                               'Количество продаж ( с учетом возврат и отмен)', 'Цена розничная с учетом согласованной скидки', 
                                                               'Вайлдберриз реализовал Товар (Пр)', 'Комиссия WB (до спп)', 'Комиссия WB (после спп)', 'Цена', 
                                                               'К перечислению Продавцу за реализованный', 'Логистика', 'Логистика средняя', 'Общая сумма штрафов', 
                                                               'Себестоимость сумма', 'Налог', 'Выручка (минус все расходов мп)', 'Маржинальность', 'ROI %', 
                                                               'ЧП по SKU', 'Остатки на складе, шт', 'Остатки МП, шт', 'Среднее кол-во заказов в день, шт', 
                                                               'Остаток денег (wb)', 'Остаток денег (свой склад)', 'Процент выкупа, %', 'Конверсия в корзину, %', 
                                                               'Конверсия в заказ, %', 'Закончится через WB', 'Закончится через wb  / (fbs)', 'Р/К бюджет', 'ДРР'])
    len_df = 37+len(fullstats1_df)

    fullstats1_df.loc[len(fullstats1_df.index)] = ['', '', '', '', '', f"=SUM(F39:F{len_df})", f"=SUM(G39:G{len_df})", f"=SUM(H39:H{len_df})", f"=SUM(I39:I{len_df})", 
                                                   f"=SUM(J39:J{len_df})", f"=SUM(K39:K{len_df})", f"=AVERAGE(L39:L{len_df})", f"=SUM(M39:M{len_df})", f"=SUM(N39:N{len_df})", 
                                                   f"=AVERAGE(O39:O{len_df})", f"=SUM(P39:P{len_df})", f"=SUM(Q39:Q{len_df})", f"=SUM(R39:R{len_df})", f"=SUM(S39:S{len_df})", 
                                                   f"=IF(S{len_df+1}=0,0,V{len_df+1}/S{len_df+1})", f"=IF(Q{len_df+1}=0,0,V{len_df+1}/Q{len_df+1})", 
                                                   f"=M{len_df+1}-N{len_df+1}-P{len_df+1}-Q{len_df+1}-R{len_df+1}-AG{len_df+1}", f"=SUM(W39:W{len_df})", f"=SUM(X39:X{len_df})", 
                                                   f"=AVERAGE(Y39:Y{len_df})", f"=SUM(Z39:Z{len_df})", f"=SUM(AA39:AA{len_df})", f"=AVERAGE(AB39:AB{len_df})", 
                                                   f"=AVERAGE(AC39:AC{len_df})", f"=AVERAGE(AD39:AD{len_df})", "", "", f"=SUM(AG39:AG{len_df})", f"=IF(S{len_df+1}=0,0,AG{len_df+1}/S{len_df+1})"]



    lst = ['Период', 'Выкуплено, шт.', 'Выкуплено, руб.', 'Заказано, шт.', 'Возврат, шт.', 'Реклама всего, руб.', 'Штрафы, руб.', 
           'Комиссия WB', 'Логистика, руб.', 'Себестоимость выкупленного товара', 'Хранение, руб.', 'Прочие расходы, руб.', 'Налог', 'Налогооблагаемая база', 'Прибыль', 'Маржинальность','ROI %']

    df2['Дата продажи'] = df2['Дата продажи'].astype('datetime64[ns]')
    df2['Дата продажи'] = pd.to_datetime(df2['Дата продажи']).dt.normalize()

    date_min = df2['Дата продажи'].min()
    date_max = df2['Дата продажи'].max()
    

    lst_2 = [f'{date_min.strftime("%d.%m")} - {date_max.strftime("%d.%m")}', f"=G{len_df+1}", f"=H{len_df+1}", f"=F{len_df+1}", 
             f"=SUM('Сводная таблица'!J3:J{3+len(fullstats1_df)})", "=$B$25", f"=P{len_df+1}", f"=J{len_df+1}", f"=N{len_df+1}", f"=Q{len_df+1}", "=$B$24", "=$B$26", 
             f"=I{len_df+1}*$B$23/100", f"=I{len_df+1}", "=B4-B7-B8-B9-B10-B11-B12-B13-B14", f"=T{len_df+1}",  f"=U{len_df+1}"]

    fullstats2_df = pd.DataFrame(list(zip(lst, lst_2)), columns=['', ''])

    fullstats2_df.to_excel(middle_path, index=False, sheet_name='Аналитика')

    update_spreadsheet(middle_path, fullstats1_df, 1, 38, sheet_name='Аналитика')



    lst = ['Денег на складе wb', 'Остатки', 'Остатки (склад wb)', 'Налог, %', 'Хранение, руб', 'Реклама, руб', 'Прочие расходы, руб']
    lst_2 = [f'=Z{len_df+1}', f'=AA{len_df+1}', f'=W{len_df+1}', '', '','','']

    fullstats3_df = pd.DataFrame(list(zip(lst, lst_2)), columns=['', ''])
    update_spreadsheet(middle_path, fullstats3_df, 1, 20, sheet_name='Аналитика')

    lst = ['Продажа', 'К перечислению', 'Логистика за период', 'Итого к оплате.']
    lst_2 = [f'=I{len_df+1}', f'=M{len_df+1}', f'=N{len_df+1}', f'=M{len_df+1}-B24-B25-B26-N{len_df+1}']

    fullstats4_df = pd.DataFrame(list(zip(lst, lst_2)), columns=['', ''])
    update_spreadsheet(middle_path, fullstats4_df, 1, 31, sheet_name='Аналитика')

    df = pd.read_excel(middle_path)

    with ExcelWriter(middle_path, mode="a" if os.path.exists(middle_path) else "w", engine='openpyxl') as writer:
        marketing_df.to_excel(writer, sheet_name="Затраты на РК", index=False)
        price_df.to_excel(writer, sheet_name="Себестоимость", index=False)
        sales_df.to_excel(writer, sheet_name="Продажи", index=False)
        fullstats_svodnaya.to_excel(writer, sheet_name="Сводная таблица", index=False)





    # Работаем с форматированием таблицы

    workbook = op.load_workbook(filename=middle_path, read_only=False)
    sheet = workbook.active
    marketing_sheet = workbook['Затраты на РК']
    fullstats_svodnaya_sheet = workbook['Сводная таблица']
    price_sheet = workbook['Себестоимость']
    sales = workbook['Продажи']

    sheet.column_dimensions["A"].width = 35
    sheet.column_dimensions["B"].width = 35
    sheet.column_dimensions["C"].width = 35
    sheet.column_dimensions["D"].width = 35
    sheet.column_dimensions["E"].width = 20
    sheet.column_dimensions["F"].width = 20
    sheet.column_dimensions["G"].width = 20
    sheet.column_dimensions["H"].width = 30
    sheet.column_dimensions["I"].width = 20
    sheet.column_dimensions["J"].width = 20
    sheet.column_dimensions["K"].width = 20
    sheet.column_dimensions["L"].width = 20
    sheet.column_dimensions["M"].width = 20
    sheet.column_dimensions["N"].width = 20
    sheet.column_dimensions["O"].width = 20
    sheet.column_dimensions["P"].width = 20
    sheet.column_dimensions["Q"].width = 20
    sheet.column_dimensions["R"].width = 20
    sheet.column_dimensions["S"].width = 20
    sheet.column_dimensions["T"].width = 20
    sheet.column_dimensions["U"].width = 20
    sheet.column_dimensions["V"].width = 20
    sheet.column_dimensions["W"].width = 20
    sheet.column_dimensions["X"].width = 20
    sheet.column_dimensions["Y"].width = 20
    sheet.column_dimensions["Z"].width = 20
    sheet.column_dimensions["AA"].width = 20
    sheet.column_dimensions["AB"].width = 20
    sheet.column_dimensions["AC"].width = 20
    sheet.column_dimensions["AD"].width = 20
    sheet.column_dimensions["AE"].width = 20
    sheet.column_dimensions["AF"].width = 20
    sheet.column_dimensions["AG"].width = 20
    sheet.column_dimensions["AH"].width = 20

   
    sheet['A19'] = 'Финансы'
    sheet.merge_cells('A19:B19')
    sheet['A19'].font = Font(bold=True)
    sheet['A19'].alignment = Alignment(horizontal='center')

    sheet['A30'] = 'Итого'
    sheet.merge_cells('A30:B30')
    sheet['A30'].font = Font(bold=True)
    sheet['A30'].alignment = Alignment(horizontal='center')

    sheet['B2'].fill = PatternFill(start_color="BCBCBC", end_color="BCBCBC", fill_type = "solid")
    
    for row in sheet.iter_rows(min_row=13, min_col=2, max_row=18, max_col=2):
        for cell in row:
            cell.fill = PatternFill(start_color="b06bf7", end_color="b06bf7", fill_type = "solid")
            cell.font = Font(bold=True, color='fcfcfc')

    

    sheet['B23'].fill = PatternFill(start_color="F09A3F", end_color="F09A3F", fill_type = "solid")
    sheet['B24'].fill = PatternFill(start_color="F09A3F", end_color="F09A3F", fill_type = "solid")
    sheet['B25'].fill = PatternFill(start_color="F09A3F", end_color="F09A3F", fill_type = "solid")
    sheet['B26'].fill = PatternFill(start_color="F09A3F", end_color="F09A3F", fill_type = "solid")

    sheet['B31'].fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type = "solid")
    sheet['B32'].fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type = "solid")
    sheet['B33'].fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type = "solid")
    sheet['B34'].fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type = "solid")

    color_scale_rule = ColorScaleRule(start_type='percentile', start_value=1, start_color="00FF0000",  # красный
                                    mid_type='percentile', mid_value=50, mid_color='ffff00',
                                    end_type='percentile', end_value=99, end_color="0000FF00")  # зеленый

    sheet.conditional_formatting.add(f"S39:S{len_df+1}", color_scale_rule)
    sheet.conditional_formatting.add(f"T39:T{len_df+1}", color_scale_rule)
    sheet.conditional_formatting.add(f"U39:U{len_df+1}", color_scale_rule)
    sheet.conditional_formatting.add(f"V39:V{len_df+1}", color_scale_rule)

    for cell in sheet["38:38"]:
        cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
        cell.fill = PatternFill(start_color="666666", end_color="666666", fill_type = "solid")
        cell.font = Font(bold=True, size=10, color='fcfcfc')
        

    sheet['B14'].number_format = '#,##0.00'
    sheet['B16'].number_format = '#,##0.00' 
    sheet['B3'].number_format = '#,##0' 
    sheet['B22'].number_format = '#,##0' 
    sheet['B6'].number_format = '#,##0' 

    for cell in sheet[F"{len_df+1}:{len_df+1}"]:
        cell.fill = PatternFill(start_color="666666", end_color="666666", fill_type = "solid")
        cell.number_format = '#,##0.00' 
        cell.font = Font(bold=True, color='fcfcfc')

    for row in sheet.iter_rows(min_row=len_df+1, min_col=19, max_row=len_df+1, max_col=22):
        for cell in row:
            cell.font = Font(bold=True, color='000000')

    border1 = borders.Side(style = None, color = 'FF000000', border_style = 'thin')
    thin = Border(left = border1, right = border1, bottom = border1, top = border1)

    for row in sheet.iter_rows(min_row=13, min_col=1, max_row=18, max_col=2):
        for cell in row:
            cell.font = Font(bold=True, size=12)

    for row in sheet.iter_rows(min_row=14, min_col=3, max_row=17, max_col=5):
        for cell in row:
            cell.font = Font(bold=True, size=12)

    for row in sheet.iter_rows(min_row=38, min_col=1, max_row=len_df+1, max_col=34):
        for cell in row:
            cell.border = thin


    for row in sheet.iter_rows(min_row=30, min_col=1, max_row=34, max_col=2):
        for cell in row:
            cell.border = thin


    for row in sheet.iter_rows(min_row=2, min_col=1, max_row=18, max_col=2):
        for cell in row:
            cell.border = thin


    for row in sheet.iter_rows(min_row=20, min_col=1, max_row=27, max_col=2):
        for cell in row:
            cell.border = thin


    for row in sheet.iter_rows(min_row=39, min_col=8, max_row=len_df, max_col=34):
        for cell in row:
            cell.number_format = '#,##0.00' 

    for row in sheet.iter_rows(min_row=39, min_col=34, max_row=len_df+1, max_col=34):
        for cell in row:
            cell.number_format = '0.00%'


    for row in sheet.iter_rows(min_row=31, min_col=2, max_row=34, max_col=2):
        for cell in row:
            cell.fill = PatternFill(start_color="ffff00", end_color="ffff00", fill_type = "solid")


    for row in sheet.iter_rows(min_row=39, min_col=5, max_row=len_df, max_col=5):
        for cell in row:
            cell.fill = PatternFill(start_color="F09A3F", end_color="F09A3F", fill_type = "solid")


    for row in sheet.iter_rows(min_row=39, min_col=6, max_row=len_df, max_col=34):
        for cell in row:
            cell.fill = PatternFill(start_color="f859ff", end_color="f859ff", fill_type = "solid")
        

    for num in [7, 12, 14, 15, 16, 26, 27, 33, 34]:
        for row in sheet.iter_rows(min_row=39, min_col=num, max_row=len_df, max_col=num):
            for cell in row:
                cell.fill = PatternFill(start_color="bcfe1a", end_color="bcfe1a", fill_type = "solid")


    for row in sheet.iter_rows(min_row=39, min_col=1, max_row=21, max_col=33):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, horizontal='center')
            
    sheet.row_dimensions[38].height = 40

    for num in [20, 21]:
        for row in sheet.iter_rows(min_row=39, min_col=num, max_row=len_df+1, max_col=num):
            for cell in row:
                cell.number_format = '0%'

    # img = op.drawing.image.Image("logo.png")
    # img.height = 400
    # img.width= 400
    # img.anchor = 'D2'
    # sheet.add_image(img)

    price_sheet.column_dimensions["A"].width = 30
    price_sheet.column_dimensions["B"].width = 30
    price_sheet.column_dimensions["C"].width = 30
    
    for cell in price_sheet["1:1"]:
        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type = "solid")

    for row in price_sheet.iter_rows(min_row=1, min_col=1, max_row=len_df-20, max_col=3):
        for cell in row:
            cell.border = thin


    marketing_sheet.column_dimensions["A"].width = 30
    marketing_sheet.column_dimensions["B"].width = 30
    marketing_sheet.column_dimensions["C"].width = 30
    for cell in marketing_sheet["1:1"]:
        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type = "solid")

    for row in marketing_sheet.iter_rows(min_row=1, min_col=1, max_row=len_df-20, max_col=3):
        for cell in row:
            cell.border = thin


    fullstats_svodnaya_sheet.column_dimensions["A"].width = 25
    fullstats_svodnaya_sheet.column_dimensions["B"].width = 20
    fullstats_svodnaya_sheet.column_dimensions["C"].width = 20
    fullstats_svodnaya_sheet.column_dimensions["D"].width = 20
    fullstats_svodnaya_sheet.column_dimensions["E"].width = 20
    fullstats_svodnaya_sheet.column_dimensions["F"].width = 20
    fullstats_svodnaya_sheet.column_dimensions["G"].width = 20
    fullstats_svodnaya_sheet.column_dimensions["H"].width = 20
    fullstats_svodnaya_sheet.column_dimensions["I"].width = 20
    fullstats_svodnaya_sheet.column_dimensions["J"].width = 20
    fullstats_svodnaya_sheet.column_dimensions["K"].width = 20
    fullstats_svodnaya_sheet.column_dimensions["L"].width = 20
    fullstats_svodnaya_sheet.column_dimensions["M"].width = 20
    fullstats_svodnaya_sheet.column_dimensions["N"].width = 20
    fullstats_svodnaya_sheet.column_dimensions["O"].width = 20
    fullstats_svodnaya_sheet.column_dimensions["P"].width = 20
    fullstats_svodnaya_sheet.column_dimensions["Q"].width = 20
    fullstats_svodnaya_sheet.column_dimensions["R"].width = 20
    fullstats_svodnaya_sheet.column_dimensions["S"].width = 20
    fullstats_svodnaya_sheet.column_dimensions["T"].width = 20
    fullstats_svodnaya_sheet.column_dimensions["U"].width = 20

    for i in range(1, 2):
        for cell in fullstats_svodnaya_sheet[f"{i}:{i}"]:
            cell.alignment = Alignment(wrap_text=True, horizontal='center')

    fullstats_svodnaya_sheet.merge_cells('B1:C1')
    fullstats_svodnaya_sheet.merge_cells('D1:E1')
    fullstats_svodnaya_sheet.merge_cells('F1:G1')
    fullstats_svodnaya_sheet.merge_cells('H1:I1')
    fullstats_svodnaya_sheet.merge_cells('J1:K1')
    fullstats_svodnaya_sheet.merge_cells('L1:M1')
    fullstats_svodnaya_sheet.merge_cells('N1:O1')
    fullstats_svodnaya_sheet.merge_cells('P1:Q1')
    fullstats_svodnaya_sheet.merge_cells('R1:S1')
    fullstats_svodnaya_sheet.merge_cells('T1:U1')

    for row in fullstats_svodnaya_sheet.iter_rows(min_row=1, min_col=1, max_row=len_df-19, max_col=21):
        for cell in row:
            cell.border = thin


    for row in fullstats_svodnaya_sheet.iter_rows(min_row=3, min_col=2, max_row=len_df-19, max_col=22):
        for cell in row:
            cell.fill = PatternFill(start_color="9FC5E8", end_color="9FC5E8", fill_type = "solid")


    for num in range(2, 22, 2):
        for row in fullstats_svodnaya_sheet.iter_rows(min_row=3, min_col=num, max_row=len_df-19, max_col=num):
            for cell in row:
                cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type = "solid")

    fullstats_svodnaya_sheet.sheet_state = 'hidden'
    sales.sheet_state = 'hidden'

    workbook.save(filename=f'/root/doc/{user_id}/Отчет_{date_min.strftime("%d.%m")}_{date_max.strftime("%d.%m")}_{user_id}.xlsx')
    date = [date_min.strftime("%d.%m"), date_max.strftime("%d.%m")]
    return date
